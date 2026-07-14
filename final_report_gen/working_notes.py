#!/usr/bin/env python3
"""
fill_template.py
-----------------
Populates Working_Notes_Template.xlsx (the file with {{PLACEHOLDER}} tags)
with data from a candidate JSON file, and writes out a finished workbook.

Edit the four paths/variables below, then run:
    python fill_template.py
"""

import json
import copy
import os
import re
import datetime
from openpyxl import load_workbook

# EDIT THESE FOR EACH RUN
TEMPLATE_PATH = r"template\WORKING_NOTES_TEMPLATE.xlsx" 
INPUT_JSON_PATH =  r"data\data_ocr.json"          
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")                              
CASE_ID = "TDCS-XXX"                   


DASH = "-"
TEMPLATE_ROW = 4  # row in the template that holds {{DOC_LABEL}}, {{NAME}}, ...


def parse_date(value):
    if not value or not isinstance(value, str):
        return DASH
    value = value.strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(value, fmt)
        except ValueError:
            continue
    return value


def val(x):
    if x is None or (isinstance(x, str) and not x.strip()):
        return DASH
    return x


def row_cv(doc):
    addr = (doc.get("current_address") or {}).get("complete_address")
    return {"label": "Curriculum-Vitae", "name": val(doc.get("full_name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")),
            "address": val(addr)}


def rows_employment(doc):
    rows = []
    for cert in (doc.get("certificates") or []):
        company = cert.get("company_name") or "Employer"
        rows.append({"label": f"Employment Certificate ({company})",
                      "name": val(cert.get("candidate_name")), "dob": DASH, "pob": DASH,
                      "father": DASH, "mother": DASH, "address": DASH})
    return rows


def row_birth_certificate(doc):
    return {"label": "Birth Certificate", "name": val(doc.get("child_name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": val(doc.get("place_of_birth")),
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")),
            "address": val(doc.get("parents_address"))}


def row_tenth(doc):
    return {"label": "Education Certificate (Standard 10th)", "name": val(doc.get("student_name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")), "address": DASH}


def row_twelfth(doc):
    return {"label": "Education Certificate (Standard 12th)", "name": val(doc.get("student_name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")), "address": DASH}


def row_passport(doc):
    given_name = (doc.get("given_name") or "").strip()
    surname = (doc.get("surname") or "").strip()
    name = " ".join(part for part in [given_name, surname] if part)
    return {"label": "Passport", "name": val(name), "dob": parse_date(doc.get("date_of_birth")),
            "pob": val(doc.get("place_of_birth")), "father": val(doc.get("father_name")),
            "mother": val(doc.get("mother_name")), "address": val(doc.get("address"))}


def row_aadhaar(doc):
    return {"label": "Aadhaar Card", "name": val(doc.get("name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")),
            "address": val(doc.get("address"))}


def row_pan(doc):
    return {"label": "PAN Card", "name": val(doc.get("name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": DASH, "address": DASH}


def row_driving_licence(doc):
    return {"label": "Driving Licence", "name": val(doc.get("name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": DASH, "address": val(doc.get("address"))}


def row_voter_id(doc):
    return {"label": "Voter ID", "name": val(doc.get("name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": val(doc.get("mother_name")),
            "address": val(doc.get("address"))}


def row_marriage(doc):
    return {"label": "Marriage Certificate", "name": val(doc.get("husband_name") or doc.get("name")),
            "dob": DASH, "pob": DASH, "father": DASH, "mother": DASH, "address": val(doc.get("address"))}


def row_internship(doc):
    label = f"Internship Certificate ({val(doc.get('company_name'))})" if doc.get("company_name") \
        else "Internship Certificate"
    return {"label": label, "name": val(doc.get("candidate_name")), "dob": DASH, "pob": DASH,
            "father": DASH, "mother": DASH, "address": DASH}


def row_transfer_certificate(doc):
    return {"label": "Transfer Certificate", "name": val(doc.get("student_name")),
            "dob": parse_date(doc.get("date_of_birth")), "pob": DASH,
            "father": val(doc.get("father_name")), "mother": DASH, "address": DASH}

_NAME_KEYS = ("name", "full_name", "candidate_name", "student_name",
              "child_name", "employee_name", "applicant_name")
_ADDRESS_KEYS = ("address", "complete_address", "permanent_address",
                  "current_address", "parents_address", "registered_address")


def _first_present(doc, keys):
    for k in keys:
        v = doc.get(k)
        if isinstance(v, dict):
            # nested address-style object -> try its complete_address/address_line
            v = v.get("complete_address") or v.get("address_line")
        if v:
            return v
    return None


def generic_row(label, doc):
    return {
        "label": label,
        "name": val(_first_present(doc, _NAME_KEYS)),
        "dob": parse_date(doc.get("date_of_birth")),
        "pob": val(doc.get("place_of_birth")),
        "father": val(doc.get("father_name")),
        "mother": val(doc.get("mother_name")),
        "address": val(_first_present(doc, _ADDRESS_KEYS)),
    }


def generic_handler(key, value):
    """Handles any top-level JSON key that isn't in KNOWN_HANDLERS below.
    Copes with: a plain dict, a list of dicts, or a dict that wraps a list
    under a sub-key like 'certificates' / 'records' / 'items'."""
    label_base = key.replace("_", " ").title()

    if isinstance(value, dict):
        for subkey in ("certificates", "records", "items", "list", "entries", "details"):
            nested = value.get(subkey)
            if isinstance(nested, list):
                return generic_handler(key, nested)
        return [generic_row(label_base, value)]

    if isinstance(value, list):
        rows = []
        for item in value:
            if not isinstance(item, dict):
                continue
            distinguishing = (item.get("company_name") or item.get("institution")
                               or item.get("issuer") or item.get("school_name"))
            label = f"{label_base} ({distinguishing})" if distinguishing else label_base
            rows.append(generic_row(label, item))
        return rows

    return []


KNOWN_HANDLERS = [
    ("cv_bgv", row_cv),
    ("employment_certificate", rows_employment),
    ("birth_certificate", row_birth_certificate),
    ("tenth_marksheet", row_tenth),
    ("twelfth_marksheet", row_twelfth),
    ("transfer_certificate", row_transfer_certificate),
    ("internship_certificate", row_internship),
    ("passport", row_passport),
    ("aadhaar_card", row_aadhaar),
    ("pan_card", row_pan),
    ("driving_licence", row_driving_licence),
    ("voter_id", row_voter_id),
    ("marriage", row_marriage),
]
KNOWN_KEYS = {key for key, _ in KNOWN_HANDLERS}


def _is_empty(value):
    """True for None, {}, [], '' — i.e. 'this document was not provided'."""
    if value is None:
        return True
    if isinstance(value, (dict, list, str)) and len(value) == 0:
        return True
    return False


def build_rows(data):
    """Walks every key actually present in the JSON. Documents that are
    null/missing/empty are simply skipped — nothing is assumed compulsory.
    Known document types use their tailored extractor; any other key
    (unexpected/new document types) falls back to generic_handler so it
    still produces a row instead of being silently ignored."""
    rows = []

    # 1) Known document types, in a stable/expected order, only if present.
    for key, handler in KNOWN_HANDLERS:
        doc = data.get(key)
        if _is_empty(doc):
            continue
        if isinstance(doc, dict):
            result = handler(doc)
        else:
            # Value came in as a list even though we expected a dict (or
            # vice versa) — fall back to the generic handler rather than
            # crashing on a shape mismatch.
            result = generic_handler(key, doc)
        if result is None:
            continue
        rows.extend(result) if isinstance(result, list) else rows.append(result)

    # 2) Anything else in the JSON that isn't one of the known keys —
    #    e.g. a document type added later that this script doesn't know
    #    about yet — still gets picked up automatically.
    for key in sorted(data.keys()):
        if key in KNOWN_KEYS:
            continue
        doc = data.get(key)
        if _is_empty(doc):
            continue
        rows.extend(generic_handler(key, doc))

    return rows


def candidate_name(data):
    for key, field in [("cv_bgv", "full_name"), ("birth_certificate", "child_name")]:
        doc = data.get(key)
        if doc and doc.get(field):
            return doc.get(field)
    passport = data.get("passport")
    if passport:
        given_name = (passport.get("given_name") or "").strip()
        surname = (passport.get("surname") or "").strip()
        combined = " ".join(part for part in [given_name, surname] if part)
        if combined:
            return combined
    return "CANDIDATE"


def fill_placeholders(text, mapping):
    if not isinstance(text, str):
        return text
    for key, value in mapping.items():
        text = text.replace(key, value)
    # If a {{CASE_ID}} placeholder is left dangling (no case id supplied),
    # clean up the surrounding " (...)" too.
    text = text.replace(" ({{CASE_ID}})", "")
    return text


def safe_filename_part(text):
    """Strip characters that are illegal in Windows/Mac/Linux filenames."""
    text = re.sub(r'[\\/:*?"<>|]', "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def main():
    with open(INPUT_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    name = candidate_name(data)
    header_map = {
        "{{CANDIDATE_NAME}}": name,
        "{{CANDIDATE_NAME_UPPER}}": name.upper(),
        "{{CASE_ID}}": CASE_ID,
    }

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # --- Fill header placeholders (rows 1-3) ---
    for coord in ["A1", "A2", "A3"]:
        ws[coord].value = fill_placeholders(ws[coord].value, header_map)

    # --- Capture the template row's style before we touch anything ---
    cols = ["A", "B", "C", "D", "E", "F", "G"]
    template_styles = {
        col: (copy.copy(ws[f"{col}{TEMPLATE_ROW}"].font),
              copy.copy(ws[f"{col}{TEMPLATE_ROW}"].alignment),
              copy.copy(ws[f"{col}{TEMPLATE_ROW}"].border))
        for col in cols
    }
    template_row_height = ws.row_dimensions[TEMPLATE_ROW].height

    rows = build_rows(data)
    if not rows:
        rows = [{"label": "No documents found in source data", "name": DASH,
                 "dob": DASH, "pob": DASH, "father": DASH, "mother": DASH, "address": DASH}]

    # --- Write first document into the existing template row ---
    def write_row(r, row_data):
        values = {
            "A": row_data["label"], "B": row_data["name"], "C": row_data["dob"],
            "D": row_data["pob"], "E": row_data["father"], "F": row_data["mother"],
            "G": row_data["address"],
        }
        for col in cols:
            cell = ws[f"{col}{r}"]
            cell.value = values[col]
            font, alignment, border = template_styles[col]
            cell.font = copy.copy(font)
            cell.alignment = copy.copy(alignment)
            cell.border = copy.copy(border)
            if col == "C" and isinstance(values[col], datetime.datetime):
                cell.number_format = "dd/mm/yyyy;@"
        ws.row_dimensions[r].height = template_row_height

    write_row(TEMPLATE_ROW, rows[0])

    # --- Insert + write a cloned row for every additional document ---
    for i, row_data in enumerate(rows[1:], start=1):
        r = TEMPLATE_ROW + i
        ws.insert_rows(r)
        write_row(r, row_data)

    # --- Build an output filename that starts with the client's name ---
    name_part = safe_filename_part(name) or "Candidate"
    filename = f"{name_part} - Working Notes"
    if CASE_ID:
        filename += f" ({safe_filename_part(CASE_ID)})"
    filename += ".xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb.save(output_path)
    print(f"Wrote {len(rows)} document rows to {output_path}")


if __name__ == "__main__":
    main()